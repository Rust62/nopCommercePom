import pytest
from openpyxl.styles import PatternFill
from Pages.Login_Page import LoginPage
from Tests.test_base import BaseTest
from read_write_excel import get_count_rows , read_data , write_data , fill_cell

path = "C:\\Users\\rusgl\\PycharmProjects\\nopCommercePOM\\Datafile.xlsx"
pattern1 = PatternFill ( 'solid' , fgColor='F50707' )
pattern2 = PatternFill ( 'solid' , fgColor='71FF33' )
Title = "Dashboard / nopCommerce administration"


class TestLoginFunctionality ( BaseTest ):

    def test_login_functionality(self):

        rows = get_count_rows ( path , "login" )
        for row in range ( 2 , rows + 1 ):
            username = read_data ( path , "login" , row , 1 )
            password = read_data ( path , "login" , row , 2 )
            self.loginPage = LoginPage ( self.driver )
            self.loginPage.do_login ( username , password )

            self.title = self.loginPage.get_login_page_title ()
            print ( self.title )
            if self.title == Title:
                print ( "Test passed" )
                write_data ( path , "login" , row , 3 , "test passed" )
                fill_cell ( path , 'login' , row , 3 , pattern2 )
            else:
                print ( "test failed" )
                write_data ( path , "login" , row , 3 , "test failed" )
                fill_cell ( path , 'login' , row , 3 , pattern1 )
