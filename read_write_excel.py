from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def get_count_rows(file , sheet_name):
    wb = load_workbook ( file )
    sheet = wb[sheet_name]
    return sheet.max_row


def get_count_columns(file , sheet_name):
    wb = load_workbook ( file )
    sheet = wb[sheet_name]
    return sheet.max_column


def read_data(file , sheet_name , row , column):
    wb = load_workbook ( file )
    sheet = wb[sheet_name]
    return sheet.cell ( row , column ).value


def write_data(file , sheet_name , row , column , data):
    wb = load_workbook ( file )
    sheet = wb[sheet_name]
    sheet.cell ( row , column ).value = data
    wb.save ( file )


def fill_cell(file , sheet_name , row , column, pattern):
    wb = load_workbook ( file )
    sheet = wb[sheet_name]
    sheet.cell ( row , column ).fill = pattern
    wb.save ( file )



    # def get_data_from_exl(self, xl_file):
    #     self.wb = load_workbook(xl_file)
    #     # print(wb.sheetnames)
    #     # print(wb.active.title)
    #     self.sheet = self.wb['active']
    #     self.rows = self.sheet.max_row
    #     self.columns = self.sheet.max_column
    #     # print ( rows , columns )
    #
    #     for row in range(2, self.rows + 1):
    #         email = self.sheet.cell(row, 1).value
    #         if row == 3:
    #             break
    # print ( email , password )

# email = sheet.cell ( 2 , 1 ).value
# password = sheet.cell ( 2 , 2 ).value


# sheet.cell(3, 1, value="rg@gmail.com")
# sheet.cell(3, 1).fill=PatternFill('solid', fgColor='71FF33')
# sheet.cell(3, 2, value="admin1")
# sheet.cell(3, 2).fill = PatternFill('solid', fgColor='F50707')
# wb.save("Datafile2.xlsx")

# print ( sheet['A2'].value )
# print ( sheet.cell ( row=2 , column=2 ).value )
